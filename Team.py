import os
import tabulate
from fpdf import FPDF
import gender_guesser.detector as gender
import re
from openpyxl import load_workbook


L_CEL = 40
H_CEL_INT = 10
H_CEL = 15
H_CEL_PL = 5


class Team:
    def __init__(self, team_name, resp_name, tel, mail):
        self.team_name = team_name
        self.resp_name = resp_name
        self.tel = tel
        self.mail = mail
        self.players = list()

    def add_player(self, player):
        self.players.append(player)

    def to_string(self):
        print(f"TEAM: {self.team_name}\n"
              f"RESPONSABILE: {self.resp_name}\n"
              f"NUMERO DI TELEFONO: {self.tel}\n"
              f"E-MAIL: {self.mail}\n"
              f"GIOCATORI")
        list_dict = []

        for p in self.players:
            list_dict.append(p.to_dict())

        print(tabulate.tabulate(list_dict, headers="keys", tablefmt='rst'))
        print("\n\n")


    def determine_gender(self, name):

        d = gender.Detector()
        guessed_gender = d.get_gender(name)
        # Mappa il risultato della libreria al formato desiderato
        if guessed_gender in ["male", "mostly_male"]:
            print(f"Nome: {name} - Risultato: {guessed_gender}")
            return "M"
        elif guessed_gender in ["female", "mostly_female"]:
            print(f"Nome: {name} - Risultato: {guessed_gender}")
            return "F"

        name = name.split()[0].capitalize()  # Prendi solo il primo nome e capitalizza
        if name.endswith("o"):
            return "M"
        elif name.endswith("a"):
            return "F"
        else:
            return " "

    def create_team_folder(self, tournament):
        os.mkdir(f"{tournament}/{self.team_name}")
        os.mkdir(f"{tournament}/{self.team_name}/AUTH")
        os.mkdir(f"{tournament}/{self.team_name}/CM")

        pdf = FPDF("L", "mm", "A4")
        pdf.set_auto_page_break(auto=False)  # Disabilita i salti automatici di pagina
        pdf.add_page()
        pdf.set_font("Times", size=9)

        # Regola i margini e larghezza della linea
        #pdf.set_margins(left=5, top=5, right=5)
        #pdf.set_line_width(0.2)

        # Aggiungi l'intestazione con un layout migliorato
        label_width = 50  # Larghezza fissa per le etichette
        value_width = 140  # Larghezza fissa per i valori
        line_height = 10  # Altezza delle righe

        # TEAM
        pdf.set_font("Times", style="B", size=10)
        pdf.cell(label_width, line_height, txt="TEAM:", ln=0, align="L")
        pdf.set_font("Times", size=10)
        pdf.cell(value_width, line_height, txt=self.team_name, ln=1, align="L")

        # RESPONSABILE
        pdf.set_font("Times", style="B", size=10)
        pdf.cell(label_width, line_height, txt="RESPONSABILE:", ln=0, align="L")
        pdf.set_font("Times", size=10)
        pdf.cell(value_width, line_height, txt=self.resp_name, ln=1, align="L")

        # NUMERO DI TELEFONO
        pdf.set_font("Times", style="B", size=10)
        pdf.cell(label_width, line_height, txt="NUMERO DI TELEFONO:", ln=0, align="L")
        pdf.set_font("Times", size=10)
        pdf.cell(value_width, line_height, txt=str(self.tel), ln=1, align="L")

        # E-MAIL
        pdf.set_font("Times", style="B", size=10)
        pdf.cell(label_width, line_height, txt="E-MAIL:", ln=0, align="L")
        pdf.set_font("Times", size=10)
        pdf.cell(value_width, line_height, txt=self.mail, ln=1, align="L")

        # Intestazioni della tabella
        pdf.set_font("Times", style="B", size=9)
        pdf.set_line_width(0.3)
        headers = ["NOME", "COGNOME", "DATA DI NASCITA", "LUOGO DI NASCITA", "CARTA IDENTITA'", "CATEGORIA",
                   "MAGLIETTA"]
        #col_width = 35
        for header in headers:
            pdf.cell(L_CEL, 10, txt=header, border='B')
        pdf.cell(0, 10, ln=1)

        wb = load_workbook("modulo_free_sport.xlsx")
        ws = wb.active

        ws["B2"] = f"{self.team_name} - {tournament}"
        start_row = 11
        for i, player in enumerate(self.players):
            row = start_row + i
            ws[f"C{row}"] = player.surname
            ws[f"D{row}"] = player.name
            ws[f"E{row}"] = ws[f"E{row}"] = "M" if tournament == "Calcetto" else self.determine_gender(player.name)
            ws[f"F{row}"] = player.date
            ws[f"G{row}"] = player.place

        # Salva il file con il nome della squadra
        output_path = os.path.join(f"{tournament}/{self.team_name}/", f"{self.team_name}.xlsx")
        wb.save(output_path)
        print(f"File creato: {output_path}")

        # Riga dei giocatori
        pdf.set_font("Times", size=8)
        for player in self.players:
            pdf.cell(L_CEL, H_CEL_PL, txt=player.name, border='B')
            pdf.cell(L_CEL, H_CEL_PL, txt=player.surname, border='B')
            pdf.cell(L_CEL, H_CEL_PL, txt=player.date, border='B')
            pdf.cell(L_CEL, H_CEL_PL, txt=player.place, border='B')
            pdf.cell(L_CEL, H_CEL_PL, txt=player.ci, border='B')
            pdf.cell(L_CEL, H_CEL_PL, txt=player.cat if player.cat else "Non tesserato", border='B')
            pdf.cell(L_CEL, H_CEL_PL, txt=player.shirt_size, border='B')
            pdf.cell(0, H_CEL_PL, ln=1)

        pdf.cell(70, 15, txt=" ")
        pdf.cell(0, 15, txt=" ", ln=1)
        # Disclaimer
        pdf.set_font("Times", 'B', size=13)

        disclaimers = [
            "L'organizzazione declina ogni responsabilità per eventuali danni a cose o persone di qualsiasi natura accaduti durante le partite. ",
            "Declina inoltre ogni responsabilità per eventuali danni a persone o cose causati da eventi naturali e/o atti vandalici.",
            "I partecipanti dovranno inoltre rispondere personalmente dei danni causati dal proprio comportamento.",
            "L'organizzazione si riserva il diritto di fare controlli sulle carte d'identità durante la manifestazione e se il numero indicato ",
            "e il numero presente sul documento non coincidono l'atleta verrà espulso dal torneo.",
            "Dichiaro che tutti i giocatori in distinta sono in possesso di regolare visita medica sportiva valida (sia non agonistica che agonistica).",
        ]
        for disclaimer in disclaimers:
            pdf.cell(0, 5, txt=disclaimer, ln=1)

        pdf.cell(0, H_CEL_PL, txt=f" ", ln=1)

        pdf.cell(0, H_CEL_PL, txt=f"Firma del Responsabile per accettazione di cui sopra:", ln=1)

        pdf.set_font("Times", size=10)
        pdf.cell(L_CEL, H_CEL_PL * 2, txt=f" ")
        pdf.cell(L_CEL, H_CEL_PL * 2, txt=f" ", border='B')
        pdf.cell(L_CEL, H_CEL_PL * 2, txt=f" ", border='B')

        # Salva il PDF
        pdf.output(f"{tournament}/{self.team_name}/{self.team_name}.pdf")
